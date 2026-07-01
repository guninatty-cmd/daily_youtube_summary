#!/usr/bin/env python3
"""
YouTube 영상 수집기
- 지정된 채널에서 전날 자정~당일 자정(KST) 사이 업로드된 영상(Shorts 제외) 수집
- 각 영상의 자막(스크립트)을 함께 추출
- Excel로 저장 후 Google Drive에 업로드 (OAuth 리프레시 토큰 사용)
"""

import os
import json
import datetime
import tempfile
from pathlib import Path

# 로컬 실행 시 .env 파일이 있으면 자동으로 읽어옴 (GitHub Actions에는 .env가 없으므로 그냥 무시됨)
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

import isodate
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload
from google.oauth2.credentials import Credentials
from google.auth.transport.requests import Request
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from youtube_transcript_api import YouTubeTranscriptApi
from youtube_transcript_api._errors import (
    TranscriptsDisabled,
    NoTranscriptFound,
    VideoUnavailable,
)

# ──────────────────────────────────────────────
# 환경변수 (GitHub Secrets에서 주입)
# ──────────────────────────────────────────────
YOUTUBE_API_KEY = os.environ["YOUTUBE_API_KEY"]
GDRIVE_CLIENT_ID = os.environ["GDRIVE_CLIENT_ID"]
GDRIVE_CLIENT_SECRET = os.environ["GDRIVE_CLIENT_SECRET"]
GDRIVE_REFRESH_TOKEN = os.environ["GDRIVE_REFRESH_TOKEN"]
GDRIVE_FOLDER_ID = os.environ.get("GDRIVE_FOLDER_ID", "")

# ──────────────────────────────────────────────
# 수집 대상 채널 핸들 (@ 제외)
# ──────────────────────────────────────────────
CHANNEL_HANDLES = [
    "snowballlabs",
    "sosumonkey",
    "singlefire",
    "김단테",
    "오늘도미국주식",
    "MK_Invest",
    "waltechman",
    "moneydo",
    "hs_academy",
    "daishintv",
    "t3chfeed",
    "hkglobalmarket",
    "supe-tv",
]

KST = datetime.timezone(datetime.timedelta(hours=9))

# Excel 셀 문자 수 한도(32767)에 안전 마진을 둔 값
MAX_CELL_CHARS = 32000

# youtube-transcript-api >=1.0 부터 인스턴스 기반 API로 변경됨
_YTT_API = YouTubeTranscriptApi()

# ──────────────────────────────────────────────
# API 클라이언트
# ──────────────────────────────────────────────
def get_youtube():
    return build("youtube", "v3", developerKey=YOUTUBE_API_KEY)


def get_drive():
    creds = Credentials(
        token=None,
        refresh_token=GDRIVE_REFRESH_TOKEN,
        client_id=GDRIVE_CLIENT_ID,
        client_secret=GDRIVE_CLIENT_SECRET,
        token_uri="https://oauth2.googleapis.com/token",
        scopes=["https://www.googleapis.com/auth/drive"],
    )
    creds.refresh(Request())
    return build("drive", "v3", credentials=creds)


# ──────────────────────────────────────────────
# YouTube 헬퍼
# ──────────────────────────────────────────────
def get_channel_info(yt, handle: str):
    resp = yt.channels().list(
        part="id,snippet,contentDetails",
        forHandle=handle,
    ).execute()

    items = resp.get("items", [])
    if not items:
        print(f"  채널 없음: @{handle}")
        return None, None, None

    ch = items[0]
    return (
        ch["id"],
        ch["snippet"]["title"],
        ch["contentDetails"]["relatedPlaylists"]["uploads"],
    )


def fetch_recent_video_ids(yt, playlist_id: str, since: datetime.datetime, until: datetime.datetime):
    results = []
    page_token = None

    while True:
        resp = yt.playlistItems().list(
            part="contentDetails,snippet",
            playlistId=playlist_id,
            maxResults=50,
            pageToken=page_token,
        ).execute()

        items = resp.get("items", [])
        if not items:
            break

        done = False
        for item in items:
            pub_str = item["contentDetails"].get("videoPublishedAt")
            if not pub_str:
                continue
            pub = datetime.datetime.fromisoformat(pub_str.replace("Z", "+00:00"))

            if pub < since:
                done = True
                break
            if pub < until:
                results.append({
                    "id": item["contentDetails"]["videoId"],
                    "title": item["snippet"]["title"],
                    "published_at": pub,
                })

        if done:
            break
        page_token = resp.get("nextPageToken")
        if not page_token:
            break

    return results


def get_durations(yt, video_ids: list) -> dict:
    durations = {}
    for i in range(0, len(video_ids), 50):
        batch = video_ids[i : i + 50]
        resp = yt.videos().list(
            part="contentDetails",
            id=",".join(batch),
        ).execute()
        for item in resp.get("items", []):
            raw = item["contentDetails"]["duration"]
            try:
                secs = int(isodate.parse_duration(raw).total_seconds())
            except Exception:
                secs = 0
            durations[item["id"]] = secs
    return durations


def is_short(duration_secs: int) -> bool:
    return duration_secs <= 100


# ──────────────────────────────────────────────
# 자막(스크립트) 추출
# ──────────────────────────────────────────────
def get_transcript(video_id: str) -> str:
    """
    자막을 한국어 우선(없으면 영어, 그것도 없으면 사용 가능한 아무 자막)으로 가져와
    하나의 문자열로 합쳐 반환한다. 자막이 없으면 안내 문구를 반환한다.
    """
    try:
        transcript_list = _YTT_API.list(video_id)
        try:
            transcript = transcript_list.find_transcript(["ko", "en"])
        except NoTranscriptFound:
            transcript = next(iter(transcript_list))

        fetched = transcript.fetch()
        text = " ".join(seg.text.replace("\n", " ").strip() for seg in fetched if seg.text)
        text = " ".join(text.split())  # 중복 공백 정리

        if len(text) > MAX_CELL_CHARS:
            text = text[:MAX_CELL_CHARS] + " ...(생략)"

        return text or "(자막 없음)"

    except (TranscriptsDisabled, NoTranscriptFound, VideoUnavailable):
        return "(자막 없음)"
    except Exception as e:
        print(f"    [자막 추출 실패] {video_id}: {e}")
        return "(자막 추출 실패)"


# ──────────────────────────────────────────────
# 수집 메인
# ──────────────────────────────────────────────
def collect():
    yt = get_youtube()
    now_kst = datetime.datetime.now(KST)

    today_midnight = now_kst.replace(hour=0, minute=0, second=0, microsecond=0)
    yesterday_midnight = today_midnight - datetime.timedelta(days=1)

    since = yesterday_midnight.astimezone(datetime.timezone.utc)
    until = today_midnight.astimezone(datetime.timezone.utc)
    date_label = yesterday_midnight.strftime("%Y-%m-%d")

    print(f"수집 기간(KST): {yesterday_midnight.strftime('%Y-%m-%d 00:00')} ~ {today_midnight.strftime('%Y-%m-%d 00:00')}")
    print(f"(UTC: {since.strftime('%Y-%m-%d %H:%M')} ~ {until.strftime('%Y-%m-%d %H:%M')})\n")

    all_videos = []

    for handle in CHANNEL_HANDLES:
        print(f"@{handle}")
        _, ch_name, playlist_id = get_channel_info(yt, handle)
        if not playlist_id:
            continue

        videos = fetch_recent_video_ids(yt, playlist_id, since, until)
        if not videos:
            print("  -> 영상 없음")
            continue

        durations = get_durations(yt, [v["id"] for v in videos])

        for v in videos:
            secs = durations.get(v["id"], 0)
            if is_short(secs):
                print(f"  [Shorts 제외] {v['title'][:45]}")
                continue

            pub_kst = v["published_at"].astimezone(KST)
            print(f"  + {v['title'][:55]}")
            print("    자막 추출 중...")
            script_text = get_transcript(v["id"])

            all_videos.append({
                "channel_name": ch_name,
                "channel_handle": f"@{handle}",
                "title": v["title"],
                "url": f"https://www.youtube.com/watch?v={v['id']}",
                "published_kst": pub_kst.strftime("%Y-%m-%d %H:%M:%S"),
                "duration_sec": secs,
                "script": script_text,
            })

    print(f"\n총 {len(all_videos)}개 영상 수집 완료")
    return all_videos, date_label


# ──────────────────────────────────────────────
# Excel
# ──────────────────────────────────────────────
def build_excel(videos: list, date_label: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = date_label

    h_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    h_font = Font(bold=True, color="FFFFFF", size=11)
    h_align = Alignment(horizontal="center", vertical="center")

    headers = ["채널명", "채널 핸들", "영상 제목", "링크", "업로드 시간 (KST)", "재생시간(초)", "스크립트(자막)"]
    col_widths = [22, 18, 62, 48, 22, 12, 80]

    for c, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = h_fill
        cell.font = h_font
        cell.alignment = h_align
        ws.column_dimensions[get_column_letter(c)].width = w

    ws.row_dimensions[1].height = 28

    even_fill = PatternFill(start_color="EBF3FB", end_color="EBF3FB", fill_type="solid")
    link_font = Font(color="0563C1", underline="single")
    wrap_align = Alignment(vertical="top", wrap_text=True)

    for r, v in enumerate(videos, 2):
        ws.cell(r, 1, v["channel_name"])
        ws.cell(r, 2, v["channel_handle"])
        ws.cell(r, 3, v["title"])
        cell_url = ws.cell(r, 4, v["url"])
        cell_url.font = link_font
        ws.cell(r, 5, v["published_kst"])
        ws.cell(r, 6, v["duration_sec"])
        cell_script = ws.cell(r, 7, v["script"])
        cell_script.alignment = wrap_align

        if r % 2 == 0:
            for c in range(1, 8):
                ws.cell(r, c).fill = even_fill

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:G{len(videos) + 1}"

    filename = f"{date_label}_유튜브영상.xlsx"
    out_path = Path(tempfile.mkdtemp()) / filename
    wb.save(out_path)
    return out_path, filename


# ──────────────────────────────────────────────
# Google Drive
# ──────────────────────────────────────────────
def get_or_create_folder(drive, name: str, parent_id: str = "") -> str:
    q = f"name='{name}' and mimeType='application/vnd.google-apps.folder' and trashed=false"
    if parent_id:
        q += f" and '{parent_id}' in parents"

    resp = drive.files().list(q=q, spaces="drive", fields="files(id)").execute()
    files = resp.get("files", [])
    if files:
        return files[0]["id"]

    meta = {"name": name, "mimeType": "application/vnd.google-apps.folder"}
    if parent_id:
        meta["parents"] = [parent_id]
    folder = drive.files().create(body=meta, fields="id").execute()
    print(f"  Drive 폴더 생성: {name}")
    return folder["id"]


def upload_to_drive(file_path: Path, filename: str, folder_id: str = ""):
    drive = get_drive()

    if not folder_id:
        folder_id = get_or_create_folder(drive, "유튜브 영상 수집")

    meta = {"name": filename, "parents": [folder_id]}
    media = MediaFileUpload(
        str(file_path),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    f = drive.files().create(body=meta, media_body=media, fields="id,webViewLink").execute()
    print(f"  업로드 완료: {f.get('webViewLink')}")


# ──────────────────────────────────────────────
# 엔트리포인트
# ──────────────────────────────────────────────
def main():
    videos, date_label = collect()

    if not videos:
        print("수집된 영상이 없습니다. 종료.")
        return

    print("\nExcel 파일 생성 중...")
    excel_path, filename = build_excel(videos, date_label)
    print(f"  저장: {filename}")

    print("\nGoogle Drive 업로드 중...")
    try:
        upload_to_drive(excel_path, filename, GDRIVE_FOLDER_ID)
        print("\n완료!")
    except Exception as e:
        import traceback
        print("\n[Drive 업로드 실패 - 상세 에러]")
        print(traceback.format_exc())
        raise


if __name__ == "__main__":
    main()
